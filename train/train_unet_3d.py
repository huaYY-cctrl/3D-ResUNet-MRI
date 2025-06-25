import os
import torch
import numpy as np
from tqdm import tqdm # 进度条库
import torch.optim as optim
from openpyxl import Workbook
from models.UNet3DMMS import UNet3DMMS # 导入自定义的3D UNet模型
from monai.losses import DiceCELoss # 导入MONAI的Dice+CE组合损失
from torch.utils.data import DataLoader
from generators.image_label_generator import Image_Label_train, Image_Label_valid # 导入自定义数据集类


def eval(net, dataloader):
    """验证集评估函数
        参数:
            net: 待评估的模型
            dataloader: 验证集数据加载器
        返回:
            np.mean(loss_valid): 验证集平均损失
    """
    net.eval() # 设置模型为评估模式（关闭Dropout等）
    loss_valid = [] # 存储每个batch的验证损失
    with torch.no_grad(): # 关闭梯度计算，节省内存
        for image, label in dataloader:
            image = image.to(device) # 数据转移到GPU/CPU
            label = label.to(device)
            label_pred = net(image) # 模型推理
            loss = dsc_ce_loss(label_pred, label) # 计算损失
            loss_valid.append(loss.item()) # 保存损失值
    return np.mean(loss_valid) # 返回平均损失


#################################################################
# 超参数配置
epoches = 1000 # 总训练轮数
batch_size = 5 # 批量大小
init_learning_rate = 0.0005 # 初始学习率
learning_rate_patience = 8 # 学习率调整耐心值（验证损失不下降的轮数）
learning_rate_factor = 0.5 # 学习率衰减因子
modelsave_path = "..\\modelsave\\UNet\\MMS\\" # 模型保存路径
train_image_patch_folder = "D:\\PythonProject\\MMS\\train_images_patches\\" # 训练集图像块路径
train_label_patch_folder = "D:\\PythonProject\\MMS\\train_masks_patches\\" # 训练集标签块路径
validation_image_patch_folder = "D:\\PythonProject\\MMS\\validation_images_patches\\" # 验证集图像块路径
validation_label_patch_folder = "D:\\PythonProject\\MMS\\validation_masks_patches\\" # 验证集标签块路径
#################################################################

# 设备配置（优先使用GPU）
device = torch.device("cuda:0" if torch.cuda.is_available() else "cpu")

# 定义损失函数（Dice损失+交叉熵损失组合）
dsc_ce_loss = DiceCELoss(to_onehot_y=True, # 将标签转为one-hot编码（适用于多类别分割）
                         # 对模型的 原始输出（logits） 应用 Softmax 激活函数，将其转换为 类别概率分布（值域 [0, 1]，各类别概率和为 1）
                         softmax=True,  # 对模型输出应用Softmax（多类别需开启）
                         squared_pred=True) # 对预测值平方（增强平滑性）

def main():
    """主训练函数"""

    # ---------------------- 1. 训练集和验证集 ------------------
    # 获取训练图像和标签文件名列表（按字母序排序，确保一一对应）
    train_image_patch_names = sorted(os.listdir(train_image_patch_folder))
    train_label_patch_names = sorted(os.listdir(train_label_patch_folder))
    # 生成图像-标签路径对列表
    image_label_train_pairs = []
    for i in range(len(train_image_patch_names)):
        train_image_patch_path = train_image_patch_folder + train_image_patch_names[i]
        train_label_patch_path = train_label_patch_folder + train_label_patch_names[i]
        image_label_train_pairs.append([train_image_patch_path, train_label_patch_path])

    # 获取验证图像和标签文件名列表（按字母序排序，确保一一对应）
    validation_image_patch_names = sorted(os.listdir(validation_image_patch_folder))
    validation_label_patch_names = sorted(os.listdir(validation_label_patch_folder))
    image_label_valid_pairs = []
    for i in range(len(validation_image_patch_names)):
        validation_image_patch_path = validation_image_patch_folder + validation_image_patch_names[i]
        validation_label_patch_path = validation_label_patch_folder + validation_label_patch_names[i]
        image_label_valid_pairs.append([validation_image_patch_path, validation_label_patch_path])



    # ---------------------- 2. 初始化模型、优化器、学习率调度器 ----------------------
    # 创建UNet3D模型（输入通道1，输出通道14，初始特征16）
    model = UNet3DMMS(input_ch=1, output_ch=4, init_feats=16).to(device)
    # 定义优化器（Adam优化器，带权重衰减防止过拟合） 自适应学习率优化算法 为不同参数设置不同学习率
    optimizer = optim.Adam(
        model.parameters(), #这是优化器的目标，即需要更新的模型参数（权重和偏置）
        lr=init_learning_rate, #控制参数更新的步长 学习率过大可能导致训练不稳定，过小则会使收敛速度变慢
        weight_decay=0.0001,  # L2正则化强度 它会在损失函数中添加一个惩罚项，与参数的平方和成正比 用于防止过拟合
        eps=1e-5  # 防止除零的小常数
    )
    # 定义学习率调度器（验证损失不降则衰减学习率）
    scheduler = optim.lr_scheduler.ReduceLROnPlateau(
        optimizer=optimizer,
        mode="min",  # 监控指标为最小值（验证损失）
        patience=learning_rate_patience,  #patience表示 "耐心值"，即容忍多少轮（epoch）没有改进后才降低学习率
        factor=learning_rate_factor  # 衰减因子 factor是学习率衰减的倍数
    )

    # ---------------------- 3. 创建数据集和数据加载器 ----------------------
    # 训练集（带数据增强）
    trainSet = Image_Label_train(image_label_pairs=image_label_train_pairs)
    # 验证集（无数据增强）
    validSet = Image_Label_valid(image_label_pairs=image_label_valid_pairs)
    # 训练集数据加载器（打乱数据，开启多进程加载）
    trainLoader = DataLoader(dataset=trainSet,
                             batch_size=batch_size,
                             shuffle=True,
                             drop_last=True,
                             num_workers=4)
    # 验证集数据加载器（不打乱，顺序加载）
    validLoader = DataLoader(dataset=validSet,
                             batch_size=batch_size,
                             shuffle=False,
                             drop_last=True,
                             num_workers=4)

    # ---------------------- 4. 开始训练循环 ----------------------
    for epoch in range(epoches):
        print("Now is the %d epoch" % (epoch + 1))
        print("The LR now is: ", optimizer.param_groups[0]["lr"])
        model.train() # 设置模型为训练模式（开启Dropout等）
        #########################################################
        file = Workbook() # 创建一个Excel工作簿对象
        # 在工作簿中创建一个工作表，命名为'loss_train_+当前轮数'，并将其设为第一个工作表
        table = file.create_sheet('loss_train_%d' % (epoch + 1), index=0)
        loss_train = []
        #########################################################
        pbar = tqdm(trainLoader, ncols=150) # 创建进度条
        # 一个epoch有多个batch
        # 一个batch有多个patch
        for iter, (image, label) in enumerate(pbar):
            image = image.to(device)
            label = label.to(device)
            optimizer.zero_grad() # 清空梯度  # 清空优化器中保存的上一次迭代的梯度信息，防止梯度累积
            label_fake = model(image) # 前向传播 将图像输入模型，得到预测结果
            loss = dsc_ce_loss(label_fake, label) # 计算损失  # 该函数通常结合了Dice损失和交叉熵损失，适用于分割任务
            loss.backward() # 反向传播 计算损失函数相对于模型各参数的梯度
            optimizer.step() # 更新参数 根据计算得到的梯度，使用优化器更新模型参数
            # 更新进度条显示信息（当前迭代/总迭代数，当前损失值）
            pbar.set_description(f"Train process {iter + 1} / {len(trainLoader)}")
            pbar.set_postfix(dscceloss=loss.item())
            loss_train.append(loss.item())
            # 将当前批次的迭代序号和损失值写入Excel表格
            table.cell(row=iter + 1, column=1, value=iter+1) # 写入迭代序号
            table.cell(row=iter + 1, column=2, value=loss.item()) # 写入损失值
        # 计算并记录当前epoch的平均训练损失（写入Excel表格的第3列）
        table.cell(row=1, column=3, value=np.mean(loss_train))
        ###############################################################
        print("Epoch[{}]({}/{}): Loss_train: {:4f}".format(epoch,
                                                           iter+1,
                                                           len(trainLoader),
                                                           np.mean(loss_train)))
        # ---------------------- 6. 验证集评估 ----------------------
        # 使用验证集评估模型性能，返回平均验证损失值
        # eval函数通常包含前向传播过程，但不进行反向传播和参数更新
        loss_valid = eval(net=model, dataloader=validLoader)
        # 将平均验证损失值写入Excel表格的第4列（第一行）
        # 这里将验证损失与之前记录的训练平均损失（第3列）放在同一行，便于对比
        table.cell(row=1, column=4, value=loss_valid)
        file.save(modelsave_path + "ResUNet_%06d.xlsx" % (epoch + 1))
        # 更新学习率调度器（根据平均验证损失调整学习率） 学习率调整后，模型会使用更小的步长继续训练，有助于更精细地收敛到最优解
        # 提高泛化能力：适当的学习率衰减可以减少过拟合，使模型在验证集上表现更好
        # 防止训练停滞：当模型接近最优解时，较小的学习率有助于精细调整参数，避免跳过最优解
        if scheduler is not None:
            scheduler.step(loss_valid)
        # ---------------------- 7. 模型保存（每10轮保存一次） ----------------------
        if epoch % 10 == 0:
            torch.save(model.state_dict(), modelsave_path + "ResUNet_%06d.pth" % (epoch + 1))


if __name__ == '__main__':
    main()
