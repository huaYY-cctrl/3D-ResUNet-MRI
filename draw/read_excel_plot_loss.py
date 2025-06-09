import os
import openpyxl
from matplotlib import pyplot as plt

# 定义Excel文件所在的文件夹路径
excel_folder = "..\\modelsave\\UNet\\MMS\\"
# 获取文件夹中的所有文件并按名称排序
files_list = sorted(os.listdir(excel_folder))
excel_names = [] # 初始化存储Excel文件名的列表
# 遍历文件列表，筛选出所有.xlsx扩展名的文件
for i in files_list:
    if os.path.splitext(i)[1] == ".xlsx":
        excel_names.append(i)

# 初始化存储训练轮次、训练损失和验证损失的列表
epoch, loss_train, loss_valid = [], [], []
# 遍历每个Excel文件
for i in range(len(excel_names)):
    epoch.append(i + 1) # 记录当前轮次（从1开始）
    excel_path = excel_folder + excel_names[i] # 构建完整的Excel文件路径
    workbook = openpyxl.load_workbook(filename=excel_path) # 打开Excel工作簿
    sheet = workbook['loss_train_%d' % (i + 1)] # 获取指定名称的工作表（格式为'loss_train_valid_轮次号'
    loss_train.append(sheet.cell(row=1, column=3).value) # 从工作表中读取训练损失数据并添加到列表
    loss_valid.append(sheet.cell(row=1, column=4).value) # 从工作表中读取验证损失数据并添加到列表

plt.figure() # 创建一个图形对象
plt.rcParams['font.serif'] = ['Times New Roman'] # 设置字体为Times New Roman
# 绘制训练损失曲线（红色线条）
plt.plot(epoch, loss_train, "r", linewidth=1.5, label="Training Dice_CE Loss")
# 绘制验证损失曲线（蓝色线条）
plt.plot(epoch, loss_valid, "b", linewidth=1.5, label="Validation Dice_CE Loss")
plt.ylim(0.2, 2.95) # 设置y轴范围为0.2到3
plt.xlabel("Epoch") # 设置x轴标签
plt.ylabel("Loss") # 设置y轴标签
plt.legend(loc="best") # 在最佳位置显示图例
plt.title("Loss of Training & Validation") # 设置图表标题
plt.show() # 显示图表
