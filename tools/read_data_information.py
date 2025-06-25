import os
import nibabel as nib
from openpyxl import Workbook

#此文件的主要作用是批量读取医学图像数据集中的 NIfTI 格式文件，提取关键信息（如体素尺寸、图像维度等），并将这些信息整理保存到 Excel 表格中，方便后续分析和处理

# 设置NIfTI图像文件所在的文件夹路径
file_folder = "D:\\PythonProject\\MMS\\train_images_series\\"
# 获取文件夹中所有文件的名称并按字母顺序排序
file_names = sorted(os.listdir(file_folder))

# 创建一个Excel工作簿来存储数据信息
book = Workbook()
# 创建一个名为"information"的工作表，作为第一个工作表
table = book.create_sheet(title="information", index=0)
# 设置Excel表头，定义每列的含义
table.cell(row=1, column=1, value="file_name") # 文件名
table.cell(row=1, column=2, value="voxel_x") # X方向体素大小
table.cell(row=1, column=3, value="voxel_y") # Y方向体素大小
table.cell(row=1, column=4, value="voxel_z") # Z方向体素大小
table.cell(row=1, column=5, value="width") # 图像宽度（X方向体素数）
table.cell(row=1, column=6, value="height") # 图像高度（Y方向体素数）
table.cell(row=1, column=7, value="depth") # 图像深度（Z方向体素数）
table.cell(row=1, column=8, value="space_x") # X方向总空间大小（mm）
table.cell(row=1, column=9, value="space_y") # Y方向总空间大小（mm）
table.cell(row=1, column=10, value="space_z") # Z方向总空间大小（mm）

# 遍历文件夹中的每个文件
for index in range(len(file_names)):
    # 打印当前正在处理的文件名，显示进度
    print("Now is reading: ", file_names[index])
    # 在Excel表中记录完整文件名（含路径）
    table.cell(row=index + 2, column=1, value=file_folder + file_names[index])
    # 使用nibabel库加载NIfTI格式的医学图像
    file = nib.load(file_folder + file_names[index])
    # 从图像头信息中获取体素尺寸（单位通常为mm），保留前三个值（X,Y,Z方向）
    voxel = file.header.get_zooms()[:3]
    width = file.shape[0] # X方向体素数
    height = file.shape[1] # Y方向体素数
    depth = file.shape[2] # Z方向体素数
    # 计算每个方向的总空间大小（单位mm），即体素大小乘以体素数
    space_x = voxel[0] * width
    space_y = voxel[1] * height
    space_z = voxel[2] * depth
    # 将提取的信息写入Excel表格的对应单元格
    table.cell(row=index + 2, column=2, value=voxel[0])
    table.cell(row=index + 2, column=3, value=voxel[1])
    table.cell(row=index + 2, column=4, value=voxel[2])
    table.cell(row=index + 2, column=5, value=width)
    table.cell(row=index + 2, column=6, value=height)
    table.cell(row=index + 2, column=7, value=depth)
    table.cell(row=index + 2, column=8, value=space_x)
    table.cell(row=index + 2, column=9, value=space_y)
    table.cell(row=index + 2, column=10, value=space_z)
# 将所有信息保存到Excel文件中
book.save("mms_train_data_information.xlsx")
